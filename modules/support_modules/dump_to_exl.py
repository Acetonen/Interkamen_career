#!/usr/bin/env python3
"""Dump data to xlsx file."""

import os
import time
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from modules.support_modules.absolyte_path_module import AbsPath
from modules.support_modules.standart_functions import BasicFunctions


class DumpToExl(BasicFunctions):
    """Dump data to xlsx file."""
    months = [
        '', 'января', 'февраля', 'марта', 'апреля', 'мая', 'июня',
        'июля', 'августа', 'сентября', 'октября', 'ноября', 'декабря'
    ]
    blanc_drill_path = AbsPath().get_path('exl_blancs', 'drill_passport.xlsx')
    blanc_ktu_path = AbsPath().get_path('exl_blancs', 'ktu.xlsx')
    drill_pass_path = AbsPath().get_path(
        'Documents', 'Буровые_паспорта', up_root=True)
    ktu_path = AbsPath().get_path(
        'Documents', 'КТУ', up_root=True)

    @classmethod
    def _create_pass_name(cls, passport):
        """Create passport name."""
        pass_name = ("{}-{}-{} {}"
                     .format(passport.params.year,
                             passport.params.month,
                             passport.params.day,
                             int(passport.params.number)))
        return pass_name

    @classmethod
    def _normilise_barehole(cls, bareholes):
        """Normalise 5+ meters bareholes to 5 meters."""
        count = 0
        new_bareholes = {}
        for key in bareholes:
            if key >= 5:
                count += bareholes[key]
            else:
                new_bareholes[key] = bareholes[key]
        if count:
            new_bareholes[5] = count
        return new_bareholes

    def dump_drill_pass(self, passport, negab=None):
        """Dump drill passport data to blanc exl file."""
        workbook = load_workbook(self.blanc_drill_path)
        worksheet = workbook.active
        if negab:
            img = Image(AbsPath.get_path('exl_blancs', 'scheme_ng.png'))
            worksheet['F4'] = 'колличество негабаритов:'
            worksheet['K4'] = int(negab)
        else:
            img = Image(AbsPath.get_path('exl_blancs', 'scheme.png'))

        worksheet.add_image(img, 'A29')
        worksheet['K1'] = int(passport.params.number)  # Passport number.
        worksheet['J5'] = str(passport.params.day)  # Day.
        worksheet['K5'] = self.months[int(passport.params.month)]  # Month.
        worksheet['M5'] = str(passport.params.year)  # Year.
        worksheet['Q6'] = str(passport.params.horizond)  # Horizond.
        worksheet['F9'] = float(passport.params.pownder)  # Pownder.
        worksheet['K9'] = int(passport.params.d_sh)  # D_SH.
        worksheet['P9'] = int(passport.params.detonators)  # Detonators.
        # Bareholes.
        row_number = 15
        norm_bareholes = self._normilise_barehole(passport.bareholes)
        for length in norm_bareholes:
            worksheet['G' + str(row_number)] = length
            worksheet['D' + str(row_number)] = int(norm_bareholes[length])
            row_number += 1
        # Volume
        volume = round(float(passport.params.pownder) * 5 +
                       float(passport.params.d_sh) / 10, 1)
        worksheet['K27'] = volume
        # Block params.
        height = float(passport.params.block_height)
        worksheet['H25'] = height
        depth = float(passport.params.block_depth)
        worksheet['P25'] = depth
        worksheet['L25'] = round(volume / height / depth, 1)
        worksheet['M8'] = round((worksheet['L25'].value - 0.4) / int(round(
            (passport.params.block_width - 0.4) / 0.35, 0)), 3) * 1000
        # Master.
        master = super().make_name_short(str(passport.params.master))
        worksheet['J47'] = master
        # Save file.
        pass_name = self._create_pass_name(passport)
        workbook.save(os.path.join(self.drill_pass_path, pass_name) + '.xlsx')
        print(
            "\nФайл сохранен:\n",
            self.drill_pass_path + '/' + pass_name + '.xlsx'
        )

    def dump_ktu(self, tmp_rpt):
        """Dump KTU data to blanc exl file."""
        ktu = tmp_rpt.workers_showing['бух.']['КТУ']
        hours = tmp_rpt.workers_showing['бух.']['часы']
        year = tmp_rpt.status['date'].split('-')[0]
        month = (
            self.months[int(tmp_rpt.status['date'].split('-')[1][:2])][:-1]+'е'
        )
        shift = tmp_rpt.status['shift']
        brig_list = {
            'Смена 1': 'Бригадой №1',
            'Смена 2': 'Бригадой №2'
        }
        brig = brig_list[shift]

        workbook = load_workbook(self.blanc_ktu_path)
        worksheet = workbook.active
        worksheet['C4'] = brig
        worksheet['C5'] = month
        worksheet['D5'] = year
        worker_number = 1
        for worker in ktu:
            row_number = 7 + worker_number
            worksheet['A' + str(row_number)] = worker_number
            worksheet['B' + str(row_number)] = super().make_name_short(worker)
            worksheet['C' + str(row_number)] = hours[worker]
            worksheet['D' + str(row_number)] = ktu[worker]
            worker_number += 1
        # Save file.
        pass_name = '-'.join([
            year, tmp_rpt.status['date'].split('-')[1][:2], shift])
        workbook.save(os.path.join(self.ktu_path, pass_name) + '.xlsx')
        print(
            "\nФайл сохранен:\n",
            self.ktu_path + '/' + pass_name + '.xlsx'
        )
        time.sleep(3)
